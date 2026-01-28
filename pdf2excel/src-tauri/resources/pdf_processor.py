#!/usr/bin/env python3
"""
PDF 处理器 - 用于解析采购订单 PDF 并导出 Excel
通过 STDIN/STDOUT 与 Rust 通信（JSON 格式）
"""

import sys
import json
import re
from datetime import datetime

# 添加嵌入式 Python 的 site-packages 路径
import os
script_dir = os.path.dirname(os.path.abspath(__file__))
python_dir = os.path.join(script_dir, 'python')

# 尝试导入 pdfplumber，如果失败则从 embedded python 导入
try:
    import pdfplumber
except ImportError:
    # 尝试从嵌入式 Python 的 site-packages 导入
    site_packages = os.path.join(python_dir, 'Lib', 'site-packages')
    if os.path.exists(site_packages):
        sys.path.insert(0, site_packages)
    import pdfplumber

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
except ImportError:
    site_packages = os.path.join(python_dir, 'Lib', 'site-packages')
    if os.path.exists(site_packages):
        sys.path.insert(0, site_packages)
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill


def extract_pdf_info(text):
    """提取 PDF 订单信息"""
    info = {
        "order_no": "",
        "supplier_no": "",
        "supplier_name": "",
        "customer_name": "",
        "currency": "CNY"
    }
    
    # 订单号
    order_match = re.search(r'订单号[:：]\s*(\S+)', text)
    if order_match:
        info["order_no"] = order_match.group(1).strip()
    
    # 供应商编号
    supplier_no_match = re.search(r'供应商[:：]\s*(\S+)', text)
    if supplier_no_match:
        info["supplier_no"] = supplier_no_match.group(1).strip()
    
    # 供应商名称
    supplier_name_match = re.search(r'供应商名称[:：]\s*\n?([^\n]+)', text)
    if supplier_name_match:
        info["supplier_name"] = supplier_name_match.group(1).strip()
    
    # 客户名（第一行）
    lines = text.strip().split('\n')
    if lines:
        info["customer_name"] = lines[0].strip()
    
    # 货币
    currency_match = re.search(r'货币[:：]\s*(\S+)', text)
    if currency_match:
        info["currency"] = currency_match.group(1).strip()
    
    return info


def parse_pdf(path):
    """解析 PDF 文件，返回订单项列表"""
    items = []
    today = datetime.now().strftime("%Y-%m-%d")
    
    with pdfplumber.open(path) as pdf:
        full_text = ""
        for page in pdf.pages:
            text = page.extract_text()
            if text:
                full_text += text + "\n"
        
        # 提取订单信息
        info = extract_pdf_info(full_text)
        order_no = info["order_no"]
        customer_name = info["customer_name"]
        
        # 解析表格数据
        for page in pdf.pages:
            tables = page.extract_tables()
            for table in tables:
                # 查找表头行
                header_idx = -1
                for i, row in enumerate(table):
                    if row and any('零件号' in str(cell) for cell in row if cell):
                        header_idx = i
                        break
                
                if header_idx >= 0 and header_idx + 1 < len(table):
                    # 提取数据行
                    for row in table[header_idx + 1:]:
                        if not row or len(row) < 6:
                            continue
                        
                        # 清理单元格
                        row = [str(cell).replace('\n', ' ').strip() if cell else '' for cell in row]
                        
                        # 查找零件号列（包含 XX.XXX.XXXXX 格式）
                        item_code = None
                        item_code_idx = -1
                        for i, cell in enumerate(row):
                            if re.match(r'\d{2}\.\d{3}\.[\dA-Za-z]+', cell):
                                item_code = cell
                                item_code_idx = i
                                break
                        
                        if not item_code:
                            continue
                        
                        # 提取其他字段
                        line_no = row[0] if row[0] and row[0].isdigit() else ""
                        
                        # 描述通常在零件号后、数量前
                        description = ""
                        qty = ""
                        unit = "pcs"
                        price = ""
                        amount = ""
                        delivery_date = ""
                        
                        # 查找数量和价格（数字模式）
                        numeric_cells = []
                        for i, cell in enumerate(row):
                            # 匹配纯数字或带逗号的数字
                            if re.match(r'^[\d,]+\.?\d*$', cell) and i > item_code_idx:
                                numeric_cells.append((i, cell))
                        
                        # 按顺序分配：数量、价格、金额
                        if len(numeric_cells) >= 1:
                            qty = numeric_cells[0][1]
                        if len(numeric_cells) >= 2:
                            price = numeric_cells[1][1]
                        if len(numeric_cells) >= 3:
                            # 金额可能带逗号，需要清理
                            amount = numeric_cells[2][1].replace(',', '')
                        
                        # 查找日期（8位数字）
                        for cell in row:
                            if re.match(r'^\d{8}$', cell):
                                delivery_date = f"{cell[:4]}-{cell[4:6]}-{cell[6:8]}"
                                break
                        
                        # 描述 = 零件号和数量之间的内容
                        desc_parts = []
                        for i, cell in enumerate(row):
                            if i > item_code_idx and i < len(row):
                                # 如果是纯数字或单位，停止
                                if re.match(r'^\d+\.?\d*$', cell) or cell.lower() in ['pcs', '件', '个', '套']:
                                    if desc_parts:
                                        break
                                    continue
                                if cell and cell != item_code:
                                    desc_parts.append(cell)
                        description = ' '.join(desc_parts)
                        
                        if qty and price:
                            item = {
                                "id": f"{order_no}_{len(items)}",
                                "日期": today,
                                "客户名": customer_name,
                                "订单号": order_no,
                                "零件号": item_code,
                                "零件描述": description or "-",
                                "数量": qty,
                                "价格": price,
                                "金额": amount or str(float(qty) * float(price)),
                                "计划交货日期": delivery_date,
                                "订单交期": delivery_date
                            }
                            items.append(item)
    
    return {"items": items, "info": info}


def export_excel(data):
    """导出数据到 Excel"""
    path = data["path"]
    items = data["data"]
    info = data.get("info", {})
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "采购订单"
    
    # 表头
    headers = ["日期", "客户名", "订单号", "零件号", "零件描述/规格", 
               "数量", "价格（未税）", "金额（未税）", "计划交货日期", "订单交期"]
    
    # 样式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 写入表头
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    
    # 设置列宽
    ws.column_dimensions['A'].width = 12  # 日期
    ws.column_dimensions['B'].width = 20  # 客户名
    ws.column_dimensions['C'].width = 15  # 订单号
    ws.column_dimensions['D'].width = 20  # 零件号
    ws.column_dimensions['E'].width = 40  # 零件描述
    ws.column_dimensions['F'].width = 10  # 数量
    ws.column_dimensions['G'].width = 15  # 价格
    ws.column_dimensions['H'].width = 15  # 金额
    ws.column_dimensions['I'].width = 15  # 计划交货日期
    ws.column_dimensions['J'].width = 15  # 订单交期
    
    # 冻结首行
    ws.freeze_panes = 'A2'
    
    # 写入数据
    for row_idx, item in enumerate(items, 2):
        ws.cell(row=row_idx, column=1, value=item.get("日期", ""))
        ws.cell(row=row_idx, column=2, value=item.get("客户名", ""))
        ws.cell(row=row_idx, column=3, value=item.get("订单号", ""))
        ws.cell(row=row_idx, column=4, value=item.get("零件号", ""))
        ws.cell(row=row_idx, column=5, value=item.get("零件描述", ""))
        ws.cell(row=row_idx, column=6, value=float(item.get("数量", 0)))
        ws.cell(row=row_idx, column=7, value=float(item.get("价格", 0)))
        ws.cell(row=row_idx, column=8, value=float(item.get("金额", 0)))
        ws.cell(row=row_idx, column=9, value=item.get("计划交货日期", ""))
        ws.cell(row=row_idx, column=10, value=item.get("订单交期", ""))
        
        # 添加边框
        for col in range(1, 11):
            ws.cell(row=row_idx, column=col).border = thin_border
    
    # 合计行
    total_row = len(items) + 3
    total_qty = sum(float(item.get("数量", 0)) for item in items)
    total_amount = sum(float(item.get("金额", 0)) for item in items)
    
    ws.cell(row=total_row, column=5, value="合计")
    ws.cell(row=total_row, column=5).font = Font(bold=True)
    ws.cell(row=total_row, column=6, value=total_qty)
    ws.cell(row=total_row, column=8, value=total_amount)
    
    wb.save(path)
    return {"success": True}


def main():
    """主函数 - 通过 STDIN 接收命令"""
    try:
        # 读取输入
        input_data = sys.stdin.read()
        if not input_data:
            print(json.dumps({"error": "No input data"}), file=sys.stderr)
            sys.exit(1)
        
        request = json.loads(input_data)
        command = request.get("command")
        
        if command == "parse":
            result = parse_pdf(request["path"])
            print(json.dumps(result))
        
        elif command == "export":
            result = export_excel(request)
            print(json.dumps(result))
        
        else:
            print(json.dumps({"error": f"Unknown command: {command}"}), file=sys.stderr)
            sys.exit(1)
    
    except Exception as e:
        print(json.dumps({"error": str(e)}), file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
